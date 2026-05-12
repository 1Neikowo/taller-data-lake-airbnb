"""
Genera el catálogo de datos consolidado del Data Lake en Excel.
Uso: python generar_catalogo.py
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT = "catalogo_datos_datalake.xlsx"

# === HOJA 1: Catálogo Bronze ===
bronze_listings = pd.DataFrame([
    ["listing_data.csv", "listing_id", "int64", "Número entero", "Identificador", "ID único de la propiedad Airbnb"],
    ["listing_data.csv", "listing_name", "string", "Texto libre", "Descriptivo", "Nombre/título de la propiedad"],
    ["listing_data.csv", "description", "string", "HTML/Texto", "Descriptivo", "Descripción completa (contiene HTML)"],
    ["listing_data.csv", "listing_type", "string", "Categoría", "Clasificación", "Tipo de listing"],
    ["listing_data.csv", "room_type", "string", "Categoría", "Clasificación", "entire_home, private_room, hotel_room, shared_room"],
    ["listing_data.csv", "cover_photo_url", "string", "URL", "Referencia", "URL de foto de portada"],
    ["listing_data.csv", "photos_count", "int64", "Número", "Descriptivo", "Cantidad de fotos"],
    ["listing_data.csv", "photo_urls", "string", "URLs", "Referencia", "Lista de URLs de fotos"],
    ["listing_data.csv", "host_id", "int64", "Número", "Identificador", "ID del anfitrión"],
    ["listing_data.csv", "host_name", "string", "Texto", "Descriptivo", "Nombre del anfitrión"],
    ["listing_data.csv", "cohost_ids", "string", "Lista", "Identificador", "IDs de co-anfitriones"],
    ["listing_data.csv", "cohost_names", "string", "Lista", "Descriptivo", "Nombres de co-anfitriones"],
    ["listing_data.csv", "superhost", "boolean", "true/false", "Clasificación", "Si el host es Superhost"],
    ["listing_data.csv", "latitude", "float64", "Coordenada", "Geoespacial", "Latitud de la propiedad"],
    ["listing_data.csv", "longitude", "float64", "Coordenada", "Geoespacial", "Longitud de la propiedad"],
    ["listing_data.csv", "guests", "float64", "Número", "Descriptivo", "Capacidad máxima de huéspedes"],
    ["listing_data.csv", "bedrooms", "float64", "Número", "Descriptivo", "Cantidad de dormitorios"],
    ["listing_data.csv", "beds", "float64", "Número", "Descriptivo", "Cantidad de camas"],
    ["listing_data.csv", "baths", "float64", "Número", "Descriptivo", "Cantidad de baños"],
    ["listing_data.csv", "registration", "string", "Texto", "Referencia", "Número de registro/licencia"],
    ["listing_data.csv", "amenities", "string", "Lista JSON", "Descriptivo", "Amenidades disponibles"],
    ["listing_data.csv", "instant_book", "boolean", "true/false", "Clasificación", "Reserva instantánea habilitada"],
    ["listing_data.csv", "professional_management", "boolean", "true/false", "Clasificación", "Gestión profesional"],
    ["listing_data.csv", "min_nights", "int64", "Número", "Operacional", "Noches mínimas requeridas"],
    ["listing_data.csv", "cancellation_policy", "string", "Categoría", "Clasificación", "Política de cancelación"],
    ["listing_data.csv", "checkin_time", "string", "HH:MM", "Operacional", "Hora de check-in"],
    ["listing_data.csv", "checkout_time", "string", "HH:MM", "Operacional", "Hora de check-out"],
    ["listing_data.csv", "guest_favorite", "boolean", "true/false", "Clasificación", "Favorito de huéspedes"],
    ["listing_data.csv", "exact_location", "boolean", "true/false", "Geoespacial", "Ubicación exacta disponible"],
    ["listing_data.csv", "currency", "string", "ISO 4217", "Referencia", "Moneda (CLP, USD)"],
    ["listing_data.csv", "cleaning_fee", "float64", "Moneda", "Financiera", "Tarifa de limpieza"],
    ["listing_data.csv", "extra_guest_fee", "float64", "Moneda", "Financiera", "Tarifa por huésped extra"],
    ["listing_data.csv", "single_fee_structure", "boolean", "true/false", "Financiera", "Estructura de tarifa única"],
    ["listing_data.csv", "num_reviews", "int64", "Número", "Métrica", "Cantidad total de reseñas"],
    ["listing_data.csv", "rating_overall", "float64", "1.0-5.0", "Métrica", "Rating general"],
    ["listing_data.csv", "rating_accuracy", "float64", "1.0-5.0", "Métrica", "Rating de precisión"],
    ["listing_data.csv", "rating_checkin", "float64", "1.0-5.0", "Métrica", "Rating de check-in"],
    ["listing_data.csv", "rating_cleanliness", "float64", "1.0-5.0", "Métrica", "Rating de limpieza"],
    ["listing_data.csv", "rating_communication", "float64", "1.0-5.0", "Métrica", "Rating de comunicación"],
    ["listing_data.csv", "rating_location", "float64", "1.0-5.0", "Métrica", "Rating de ubicación"],
    ["listing_data.csv", "rating_value", "float64", "1.0-5.0", "Métrica", "Rating de valor"],
    ["listing_data.csv", "ttm_revenue", "float64", "USD", "Financiera", "Ingresos últimos 12 meses (USD)"],
    ["listing_data.csv", "ttm_occupancy", "float64", "0.0-1.0", "Operacional", "Ocupación últimos 12 meses"],
    ["listing_data.csv", "l90d_revenue", "float64", "USD", "Financiera", "Ingresos últimos 90 días (USD)"],
    ["listing_data.csv", "l90d_occupancy", "float64", "0.0-1.0", "Operacional", "Ocupación últimos 90 días"],
], columns=["Archivo Origen", "Columna", "Tipo de Dato", "Formato", "Clasificación", "Descripción"])

bronze_calendar = pd.DataFrame([
    ["past_calendar.csv", "listing_id", "int64", "Número entero", "Identificador / FK", "ID de la propiedad (clave de unión)"],
    ["past_calendar.csv", "date", "string", "YYYY-MM-DD", "Temporal", "Primer día del mes del registro"],
    ["past_calendar.csv", "vacant_days", "int64", "Número", "Operacional", "Días vacantes en el mes"],
    ["past_calendar.csv", "reserved_days", "int64", "Número", "Operacional", "Días reservados en el mes"],
    ["past_calendar.csv", "occupancy", "float64", "0.0-1.0", "Operacional", "Tasa de ocupación mensual"],
    ["past_calendar.csv", "revenue", "float64", "USD", "Financiera", "Ingresos del mes (USD)"],
    ["past_calendar.csv", "rate_avg", "float64", "USD", "Financiera", "Tarifa diaria promedio (USD)"],
    ["past_calendar.csv", "booked_rate_avg", "float64", "USD", "Financiera", "Tarifa promedio días reservados (USD)"],
    ["past_calendar.csv", "booking_lead_time_avg", "float64", "Días", "Operacional", "Anticipación promedio de reserva"],
    ["past_calendar.csv", "length_of_stay_avg", "float64", "Días", "Operacional", "Duración promedio de estadía"],
    ["past_calendar.csv", "min_nights_avg", "float64", "Noches", "Operacional", "Noches mínimas promedio"],
    ["past_calendar.csv", "native_booked_rate_avg", "float64", "CLP", "Financiera", "Tarifa reservados en CLP"],
    ["past_calendar.csv", "native_rate_avg", "float64", "CLP", "Financiera", "Tarifa promedio en CLP"],
    ["past_calendar.csv", "native_revenue", "float64", "CLP", "Financiera", "Ingresos del mes en CLP"],
], columns=["Archivo Origen", "Columna", "Tipo de Dato", "Formato", "Clasificación", "Descripción"])

bronze = pd.concat([bronze_listings, bronze_calendar], ignore_index=True)

# === HOJA 2: Catálogo Silver ===
silver_listings = pd.DataFrame([
    ["listings_data", "listing_id", "INTEGER", "PK", "ID único de la propiedad", "Sin cambios"],
    ["listings_data", "listing_name", "TEXT", "NOT NULL*", "Nombre de la propiedad", "Sin cambios"],
    ["listings_data", "latitude", "REAL", "Nullable", "Latitud geográfica", "Sin cambios"],
    ["listings_data", "longitude", "REAL", "Nullable", "Longitud geográfica", "Sin cambios"],
    ["listings_data", "host_name", "TEXT", "NOT NULL", "Nombre del anfitrión", "NULL → 'Desconocido'"],
    ["listings_data", "superhost", "INTEGER", "Nullable", "Si el host es Superhost (0/1)", "Sin cambios"],
    ["listings_data", "room_type", "TEXT", "NOT NULL", "Tipo de habitación", "Filtrado: solo valores válidos"],
    ["listings_data", "bedrooms", "REAL", "Nullable", "Cantidad de dormitorios", "NULL mantenido (≠ 0)"],
    ["listings_data", "beds", "REAL", "Nullable", "Cantidad de camas", "NULL mantenido (≠ 0)"],
    ["listings_data", "baths", "REAL", "Nullable", "Cantidad de baños", "NULL mantenido"],
    ["listings_data", "guests", "REAL", "Nullable", "Capacidad huéspedes", "NULL mantenido (≠ 0)"],
    ["listings_data", "currency", "TEXT", "Nullable", "Moneda (CLP)", "Filtrado valor '33' inválido"],
    ["listings_data", "num_reviews", "INTEGER", "Nullable", "Cantidad total de reseñas", "Sin cambios. NULL = sin dato"],
    ["listings_data", "rating_overall", "REAL", "Nullable", "Rating general (1.0-5.0)", "Sin cambios. NULL = sin reseñas"],
    ["listings_data", "ciudad", "TEXT", "NOT NULL", "Ciudad de origen", "DERIVADO del path de carpeta Bronze"],
], columns=["Tabla", "Columna", "Tipo SQLite", "Restricción", "Descripción", "Transformación aplicada"])

silver_calendar = pd.DataFrame([
    ["past_calendar_rates", "listing_id", "INTEGER", "FK", "ID propiedad (ref listings_data)", "Sin cambios"],
    ["past_calendar_rates", "date", "TEXT", "NOT NULL", "Fecha del mes (YYYY-MM-DD)", "Validado formato ISO"],
    ["past_calendar_rates", "vacant_days", "INTEGER", "NOT NULL", "Días vacantes", "Sin cambios"],
    ["past_calendar_rates", "reserved_days", "INTEGER", "NOT NULL", "Días reservados", "Sin cambios"],
    ["past_calendar_rates", "occupancy", "REAL", "NOT NULL", "Tasa ocupación (0.0-1.0)", "Sin cambios"],
    ["past_calendar_rates", "revenue", "REAL", "NOT NULL", "Ingresos USD", "Sin cambios"],
    ["past_calendar_rates", "rate_avg", "REAL", "NOT NULL", "Tarifa promedio USD", "Sin cambios"],
    ["past_calendar_rates", "booked_rate_avg", "REAL", "Nullable", "Tarifa reservados USD", "NULL = sin reservas"],
    ["past_calendar_rates", "booking_lead_time_avg", "REAL", "Nullable", "Anticipación reserva (días)", "NULL = sin reservas"],
    ["past_calendar_rates", "length_of_stay_avg", "REAL", "Nullable", "Duración estadía (días)", "NULL = sin reservas"],
    ["past_calendar_rates", "min_nights_avg", "REAL", "Nullable", "Noches mínimas promedio", "NULL = sin dato"],
    ["past_calendar_rates", "native_booked_rate_avg", "REAL", "Nullable", "Tarifa reservados CLP", "NULL = sin reservas"],
    ["past_calendar_rates", "native_rate_avg", "REAL", "NOT NULL", "Tarifa promedio CLP", "Sin cambios"],
    ["past_calendar_rates", "native_revenue", "REAL", "NOT NULL", "Ingresos CLP", "Sin cambios"],
    ["past_calendar_rates", "ciudad", "TEXT", "NOT NULL", "Ciudad de origen", "DERIVADO del path Bronze"],
], columns=["Tabla", "Columna", "Tipo SQLite", "Restricción", "Descripción", "Transformación aplicada"])

silver = pd.concat([silver_listings, silver_calendar], ignore_index=True)

# === HOJA 3: Catálogo de Fuentes ===
fuentes = pd.DataFrame([
    ["AirROI – Listings Data", "Datos Externos (Portal)", "AirROI Data Team", "Mensual",
     "Correo/Password AirROI", "Ficha de cada propiedad activa: nombre, ubicación, capacidad, amenidades, ratings y métricas acumuladas",
     "~300 listings por ciudad", "https://www.airroi.com/data-portal"],
    ["AirROI – Past Calendar Rates", "Datos Externos (Portal)", "AirROI Data Team", "Mensual (12 meses)",
     "Correo/Password AirROI", "Historial mensual por propiedad: ocupación, ingresos, tarifas, anticipación de reserva",
     "~300 registros/mes × 12 meses × 4 ciudades", "https://www.airroi.com/data-portal"],
    ["Dataset Vinculado (Dato Derivado)", "Procesado Internamente", "Equipo Data Analytics", "Con cada descarga",
     "Acceso interno", "Cruce entre Calendar y Listings via listing_id. Agrega domicilio a cada registro de reserva",
     "= volumen de Calendar + columnas de Listings", "Generado por ETL"],
], columns=["Nombre Fuente", "Tipo", "Data Steward", "Periodicidad", "Credenciales", "Descripción", "Volumen", "URL/Referencia"])

# === HOJA 4: Linaje de Datos ===
linaje = pd.DataFrame([
    ["listing_data.csv", "listing_id", "listings_data", "listing_id", "Filtro: solo numéricos válidos + dedup"],
    ["listing_data.csv", "listing_name", "listings_data", "listing_name", "Sin cambios"],
    ["listing_data.csv", "latitude", "listings_data", "latitude", "Sin cambios"],
    ["listing_data.csv", "longitude", "listings_data", "longitude", "Sin cambios"],
    ["listing_data.csv", "host_name", "listings_data", "host_name", "NULL → 'Desconocido'"],
    ["listing_data.csv", "superhost", "listings_data", "superhost", "Sin cambios"],
    ["listing_data.csv", "room_type", "listings_data", "room_type", "Filtro: solo entire_home, private_room, hotel_room, shared_room"],
    ["listing_data.csv", "bedrooms", "listings_data", "bedrooms", "NULL mantenido (semántico: ≠ 0)"],
    ["listing_data.csv", "beds", "listings_data", "beds", "NULL mantenido"],
    ["listing_data.csv", "baths", "listings_data", "baths", "NULL mantenido"],
    ["listing_data.csv", "guests", "listings_data", "guests", "NULL mantenido (semántico: ≠ 0)"],
    ["listing_data.csv", "currency", "listings_data", "currency", "Filtro: valor '33' eliminado con fila corrupta"],
    ["listing_data.csv", "num_reviews", "listings_data", "num_reviews", "Sin cambios. NULL mantenido"],
    ["listing_data.csv", "rating_overall", "listings_data", "rating_overall", "Sin cambios. NULL = propiedad sin reseñas"],
    ["(path carpeta)", "—", "listings_data", "ciudad", "DERIVADO: nombre carpeta Bronze → campo ciudad"],
    ["past_calendar.csv", "listing_id", "past_calendar_rates", "listing_id", "Validación numérica"],
    ["past_calendar.csv", "date", "past_calendar_rates", "date", "Validación formato YYYY-MM-DD"],
    ["past_calendar.csv", "vacant_days", "past_calendar_rates", "vacant_days", "Sin cambios"],
    ["past_calendar.csv", "reserved_days", "past_calendar_rates", "reserved_days", "Sin cambios"],
    ["past_calendar.csv", "occupancy", "past_calendar_rates", "occupancy", "Sin cambios"],
    ["past_calendar.csv", "revenue", "past_calendar_rates", "revenue", "Sin cambios"],
    ["past_calendar.csv", "rate_avg", "past_calendar_rates", "rate_avg", "Sin cambios"],
    ["past_calendar.csv", "booked_rate_avg", "past_calendar_rates", "booked_rate_avg", "NULL mantenido (meses sin reservas)"],
    ["past_calendar.csv", "booking_lead_time_avg", "past_calendar_rates", "booking_lead_time_avg", "NULL mantenido"],
    ["past_calendar.csv", "length_of_stay_avg", "past_calendar_rates", "length_of_stay_avg", "NULL mantenido"],
    ["past_calendar.csv", "min_nights_avg", "past_calendar_rates", "min_nights_avg", "NULL mantenido"],
    ["past_calendar.csv", "native_booked_rate_avg", "past_calendar_rates", "native_booked_rate_avg", "NULL mantenido"],
    ["past_calendar.csv", "native_rate_avg", "past_calendar_rates", "native_rate_avg", "Sin cambios"],
    ["past_calendar.csv", "native_revenue", "past_calendar_rates", "native_revenue", "Sin cambios"],
    ["(path carpeta)", "—", "past_calendar_rates", "ciudad", "DERIVADO: nombre carpeta Bronze → campo ciudad"],
], columns=["Origen Bronze", "Columna Bronze", "Destino Silver", "Columna Silver", "Transformación"])

# === ESCRIBIR EXCEL ===
with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
    bronze.to_excel(writer, sheet_name="Catálogo Bronze", index=False)
    silver.to_excel(writer, sheet_name="Catálogo Silver", index=False)
    fuentes.to_excel(writer, sheet_name="Catálogo Fuentes", index=False)
    linaje.to_excel(writer, sheet_name="Linaje de Datos", index=False)

# === FORMATEAR ===
wb = load_workbook(OUTPUT)
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

for ws in wb.worksheets:
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

wb.save(OUTPUT)
print(f"Catalogo generado: {OUTPUT}")
